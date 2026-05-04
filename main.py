# =====================================================
#   بوت تيليجرام لتنظيم المصاريف الشخصية
#   ✅ مجاني 100% – Google Gemini Free API
# =====================================================
import os, logging, json, io
from datetime import datetime
from collections import defaultdict

from telegram import Update
from telegram.ext import Application, MessageHandler, CommandHandler, filters, ContextTypes

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec

import arabic_reshaper
from bidi.algorithm import get_display
import google.generativeai as genai

TELEGRAM_TOKEN   = os.getenv('TELEGRAM_TOKEN')
GEMINI_API_KEY   = os.getenv('GEMINI_API_KEY')
AZURE_CLIENT_ID  = os.getenv('AZURE_CLIENT_ID')
MS_REFRESH_TOKEN = os.getenv('MS_REFRESH_TOKEN')

ONEDRIVE_FILE_PATH = 'مصاريفي.xlsx'
GRAPH_BASE = 'https://graph.microsoft.com/v1.0'
TOKEN_URL  = 'https://login.microsoftonline.com/common/oauth2/v2.0/token'
SCOPES     = ['Files.ReadWrite', 'offline_access']

logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

genai.configure(api_key=GEMINI_API_KEY)
gemini = genai.GenerativeModel('gemini-1.5-flash-latest')
_refresh_token = MS_REFRESH_TOKEN

CATEGORIES = {
    'سكن':   ['إيجار','كهرباء','غاز','مياه','صيانة','إنترنت'],
    'يومي':  ['أكل','مطعم','بقالة','قهوة','سوبرماركت'],
    'تنقل':  ['وقود','بنزين','أوبر','تاكسي','صيانة سيارة'],
    'ترفيه': ['سينما','ألعاب','سفر','فندق','هدايا'],
    'صحة':   ['دكتور','دواء','تحاليل','نظارة','صيدلية'],
    'تعليم': ['كتب','كورسات','مدرسة','جامعة'],
    'ملابس': ['ملابس','أحذية','اكسسوارات'],
    'أخرى':  ['متفرقات'],
}
PALETTE   = ['#FF6B6B','#4ECDC4','#45B7D1','#96CEB4','#FFEAA7','#DDA0DD','#98D8C8','#F0A500']
MONTHS_AR = {1:'يناير',2:'فبراير',3:'مارس',4:'أبريل',5:'مايو',6:'يونيو',
             7:'يوليو',8:'أغسطس',9:'سبتمبر',10:'أكتوبر',11:'نوفمبر',12:'ديسمبر'}

def ar(t):
    try: return get_display(arabic_reshaper.reshape(str(t)))
    except: return str(t)

# ── Microsoft Auth ────────────────────────────────
def get_ms_access_token():
    global _refresh_token
    r = requests.post(TOKEN_URL, data={
        'client_id': AZURE_CLIENT_ID, 'grant_type': 'refresh_token',
        'refresh_token': _refresh_token, 'scope': ' '.join(SCOPES),
    })
    res = r.json()
    if 'access_token' not in res:
        raise Exception(f"MS Auth Error: {res.get('error_description', res)}")
    if 'refresh_token' in res: _refresh_token = res['refresh_token']
    return res['access_token']

# ── Excel Helpers ─────────────────────────────────
def download_excel():
    token = get_ms_access_token()
    r = requests.get(f'{GRAPH_BASE}/me/drive/root:/{ONEDRIVE_FILE_PATH}:/content',
                     headers={'Authorization': f'Bearer {token}'})
    return io.BytesIO(r.content) if r.status_code == 200 else build_excel_template()

def upload_excel(wb):
    token = get_ms_access_token()
    buf   = io.BytesIO(); wb.save(buf); buf.seek(0)
    r = requests.put(
        f'{GRAPH_BASE}/me/drive/root:/{ONEDRIVE_FILE_PATH}:/content',
        headers={'Authorization': f'Bearer {token}',
                 'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'},
        data=buf.read()
    )
    if r.status_code not in (200, 201): raise Exception(f'Upload failed {r.status_code}')

def build_excel_template():
    wb = openpyxl.Workbook()
    DARK, MED, WHITE, RED, LIGHT = '1F3864','2E75B6','FFFFFF','C00000','BDD7EE'
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws = wb.active; ws.title = 'البيانات'; ws.sheet_view.rightToLeft = True
    hf = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
    hfont = Font(bold=True, color=WHITE, size=12)
    hdrs = ['التاريخ','الوقت','المبلغ','الفئة','الفئة الفرعية','ملاحظات','الشهر','السنة']
    widths = [14,10,14,15,20,30,14,8]
    for col,(h,w) in enumerate(zip(hdrs,widths),1):
        c = ws.cell(1,col,h); c.fill=hf; c.font=hfont; c.alignment=center
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 35; ws.freeze_panes = 'A2'

    wd = wb.create_sheet('الداشبورد'); wd.sheet_view.rightToLeft = True
    wd.merge_cells('A1:J1'); tc=wd['A1']
    tc.value='لوحة معلومات المصاريف الشخصية'
    tc.font=Font(bold=True,size=20,color=WHITE); tc.alignment=Alignment(horizontal='center',vertical='center')
    tc.fill=PatternFill(start_color=DARK,end_color=DARK,fill_type='solid')
    wd.row_dimensions[1].height = 50

    cat_list = list(CATEGORIES.keys())
    for cl,lb in [('A','الفئة'),('B','الإجمالي'),('C','هذا الشهر'),('D','النسبة')]:
        c=wd[f'{cl}3']; c.value=lb; c.font=Font(bold=True,color=WHITE,size=11)
        c.fill=PatternFill(start_color=MED,end_color=MED,fill_type='solid'); c.alignment=center

    for i,cat in enumerate(cat_list):
        r=4+i; fill=LIGHT if i%2==0 else WHITE
        wd.cell(r,1,cat).alignment=center
        wd.cell(r,2).value=f'=IFERROR(SUMIF(البيانات!D:D,A{r},البيانات!C:C),0)'
        wd.cell(r,2).number_format='#,##0.00'
        wd.cell(r,3).value=(f'=IFERROR(SUMPRODUCT((البيانات!D2:D10000=A{r})'
                            f'*(TEXT(البيانات!A2:A10000,"yyyy-mm")=TEXT(TODAY(),"yyyy-mm"))'
                            f'*البيانات!C2:C10000),0)')
        wd.cell(r,3).number_format='#,##0.00'
        wd.cell(r,4).value=f'=IFERROR(B{r}/SUM(B4:B{4+len(cat_list)-1}),0)'
        wd.cell(r,4).number_format='0.0%'
        for col in range(1,5):
            wd.cell(r,col).fill=PatternFill(start_color=fill,end_color=fill,fill_type='solid')
            wd.cell(r,col).alignment=center

    tr=4+len(cat_list)
    wd.cell(tr,1,'الإجمالي').font=Font(bold=True,size=12)
    wd.cell(tr,2).value=f'=SUM(B4:B{tr-1})'; wd.cell(tr,2).number_format='#,##0.00'
    wd.cell(tr,2).font=Font(bold=True,color=RED,size=12)
    wd.cell(tr,3).value=f'=SUM(C4:C{tr-1})'; wd.cell(tr,3).number_format='#,##0.00'
    for col in range(1,5):
        wd.cell(tr,col).fill=PatternFill(start_color='FCE4D6',end_color='FCE4D6',fill_type='solid')
        wd.cell(tr,col).alignment=center

    pie=PieChart(); pie.title='توزيع المصاريف'; pie.style=10; pie.width=18; pie.height=14
    pie.add_data(Reference(wd,min_col=2,min_row=3,max_row=3+len(cat_list)))
    pie.set_categories(Reference(wd,min_col=1,min_row=4,max_row=3+len(cat_list))); wd.add_chart(pie,'F2')

    bar=BarChart(); bar.type='bar'; bar.title='مصاريف الشهر'; bar.style=10; bar.width=18; bar.height=14
    bar.add_data(Reference(wd,min_col=3,min_row=3,max_row=3+len(cat_list)))
    bar.set_categories(Reference(wd,min_col=1,min_row=4,max_row=3+len(cat_list))); wd.add_chart(bar,'F18')

    for cl,w in [('A',18),('B',18),('C',18),('D',10)]:
        wd.column_dimensions[cl].width = w

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

def add_expense(expense):
    buf=download_excel(); wb=openpyxl.load_workbook(buf); ws=wb['البيانات']
    now=datetime.now(); nr=ws.max_row+1
    rf=PatternFill(start_color='EBF3FB' if nr%2==0 else 'FFFFFF',
                   end_color  ='EBF3FB' if nr%2==0 else 'FFFFFF', fill_type='solid')
    center=Alignment(horizontal='center',vertical='center')
    for col,val in enumerate([now.strftime('%Y-%m-%d'),now.strftime('%H:%M'),
                               expense['amount'],expense['category'],expense['subcategory'],
                               expense['notes'],now.strftime('%Y-%m'),now.year],1):
        c=ws.cell(row=nr,column=col,value=val); c.fill=rf; c.alignment=center
        if col==3: c.number_format='#,##0.00'; c.font=Font(bold=True,color='C00000',size=11)
    upload_excel(wb); return nr-1

def fetch_expenses():
    buf=download_excel(); wb=openpyxl.load_workbook(buf); ws=wb['البيانات']
    return [{'date':str(r[0]),'amount':float(r[2] or 0),'category':str(r[3] or 'أخرى'),
             'subcategory':str(r[4] or ''),'notes':str(r[5] or ''),'month':str(r[6] or '')}
            for r in ws.iter_rows(min_row=2,values_only=True) if r[0] and r[2]]

# ── AI Parsing (Gemini Free) ──────────────────────
def parse_expense_ai(text):
    cats = ', '.join(CATEGORIES.keys())
    prompt = f"""أنت مساعد لتحليل رسائل المصاريف باللغة العربية.
حلل الرسالة: "{text}"
الفئات: {cats}
أرجع JSON فقط بدون markdown:
إذا مصروف: {{"is_expense":true,"amount":رقم,"category":"فئة","subcategory":"وصف","notes":"ملاحظة"}}
إذا لا: {{"is_expense":false}}"""
    resp = gemini.generate_content(prompt)
    raw  = resp.text.strip().replace('```json','').replace('```','').strip()
    return json.loads(raw)

# ── Report Generation ─────────────────────────────
def build_report_image(expenses):
    now=datetime.now(); cur=now.strftime('%Y-%m')
    data=[e for e in expenses if e['month']==cur] or expenses
    total=sum(e['amount'] for e in data); count=len(data)
    avg=total/max(count,1); best=max((e['amount'] for e in data),default=0)
    ct=defaultdict(float)
    for e in data: ct[e['category']]+=e['amount']
    ct=dict(sorted(ct.items(),key=lambda x:x[1],reverse=True))

    BG,CARD,BORDER='#0D1117','#161B22','#30363D'
    TEXT,MUTED='#E6EDF3','#8B949E'
    RED,TEAL,BLUE,GOLD='#FF6B6B','#4ECDC4','#79C0FF','#FFEAA7'

    fig=plt.figure(figsize=(14,10),facecolor=BG)
    gs=gridspec.GridSpec(3,4,figure=fig,top=0.91,bottom=0.07,left=0.05,right=0.97,hspace=0.55,wspace=0.4)
    fig.text(0.5,0.955,ar(f'تقرير المصاريف | {MONTHS_AR[now.month]} {now.year}'),
             ha='center',fontsize=16,color=TEXT,fontweight='bold')
    fig.text(0.5,0.928,ar(f'تم الإنشاء: {now.strftime("%Y-%m-%d %H:%M")}'),ha='center',fontsize=9,color=MUTED)

    for i,(label,val,unit,color) in enumerate([
        ('إجمالي الشهر',f'{total:,.0f}','ج.م',RED),
        ('عدد العمليات',str(count),'عملية',BLUE),
        ('متوسط العملية',f'{avg:,.0f}','ج.م',TEAL),
        ('أكبر مصروف',f'{best:,.0f}','ج.م',GOLD)]):
        ax=fig.add_subplot(gs[0,i]); ax.set_facecolor(CARD)
        for sp in ax.spines.values(): sp.set_color(color); sp.set_linewidth(2)
        ax.set_xticks([]); ax.set_yticks([])
        ax.text(0.5,0.62,val,ha='center',va='center',fontsize=22,color=color,fontweight='bold',transform=ax.transAxes)
        ax.text(0.5,0.32,ar(unit),ha='center',va='center',fontsize=9,color=MUTED,transform=ax.transAxes)
        ax.set_title(ar(label),color=MUTED,fontsize=9,pad=8)

    ap=fig.add_subplot(gs[1,:2]); ap.set_facecolor(CARD)
    for sp in ap.spines.values(): sp.set_color(BORDER)
    ap.set_title(ar('توزيع الفئات'),color=MUTED,fontsize=9,pad=7)
    if ct:
        sizes=list(ct.values()); labels=list(ct.keys())
        colors=[PALETTE[i%len(PALETTE)] for i in range(len(labels))]
        _,_,autos=ap.pie(sizes,autopct='%1.0f%%',colors=colors,startangle=90,pctdistance=0.75,
                         wedgeprops={'edgecolor':BG,'linewidth':2,'width':0.65})
        for a in autos: a.set_color(TEXT); a.set_fontsize(8)
        ap.text(0,0.05,ar('الإجمالي'),ha='center',fontsize=8,color=MUTED)
        ap.text(0,-0.18,f'{total:,.0f}',ha='center',fontsize=12,color=RED,fontweight='bold')
        for i,(l,c) in enumerate(list(zip(labels,colors))[:6]):
            ap.text(1.08,0.88-i*0.18,'●',transform=ap.transAxes,color=c,fontsize=10)
            ap.text(1.16,0.88-i*0.18,ar(l),transform=ap.transAxes,color=TEXT,fontsize=8,va='center')

    ab=fig.add_subplot(gs[1,2:]); ab.set_facecolor(CARD)
    for sp in ab.spines.values(): sp.set_color(BORDER)
    ab.set_title(ar('المبلغ حسب الفئة'),color=MUTED,fontsize=9,pad=7)
    if ct:
        cats=list(ct.keys())[:7]; vals=[ct[c] for c in cats]
        bcols=[PALETTE[i%len(PALETTE)] for i in range(len(cats))]
        bars=ab.barh(range(len(cats)),vals,color=bcols,edgecolor='none',height=0.65)
        for bar,val in zip(bars,vals):
            ab.text(bar.get_width()+max(vals)*0.02,bar.get_y()+bar.get_height()/2,
                    f'{val:,.0f}',va='center',ha='left',color=TEXT,fontsize=8)
        ab.set_yticks(range(len(cats))); ab.set_yticklabels([ar(c) for c in cats],color=TEXT,fontsize=9)
        ab.set_xlim(0,max(vals)*1.28); ab.invert_yaxis(); ab.tick_params(axis='x',colors=MUTED,labelsize=7)

    at=fig.add_subplot(gs[2,:]); at.set_facecolor(CARD)
    for sp in at.spines.values(): sp.set_color(BORDER)
    at.set_title(ar('آخر المعاملات'),color=MUTED,fontsize=9,pad=7); at.axis('off')
    recent=sorted(data,key=lambda x:x['date'],reverse=True)[:7]
    cx=[0.02,0.15,0.30,0.50,0.72,0.96]; cat_keys=list(ct.keys())
    at.axhline(0.97,color=BORDER,lw=0.8,xmin=0.01,xmax=0.99,transform=at.transAxes)
    for j,(h,x) in enumerate(zip(['التاريخ','الفئة','التفاصيل','الملاحظات','المبلغ'],cx)):
        at.text(x,0.91,ar(h),transform=at.transAxes,color=MUTED,fontsize=9,fontweight='bold',
               ha='right' if j==4 else 'left')
    at.axhline(0.85,color=BORDER,lw=0.5,xmin=0.01,xmax=0.99,transform=at.transAxes)
    for i,e in enumerate(recent):
        y=0.76-i*0.112; col=PALETTE[cat_keys.index(e['category'])%len(PALETTE)] if e['category'] in cat_keys else TEXT
        at.text(cx[0],y,e['date'][:10],transform=at.transAxes,color=MUTED,fontsize=8)
        at.text(cx[1],y,ar(e['category']),transform=at.transAxes,color=col,fontsize=8,fontweight='bold')
        at.text(cx[2],y,ar(e['subcategory'][:18]),transform=at.transAxes,color=TEXT,fontsize=8)
        at.text(cx[3],y,ar(e['notes'][:24]),transform=at.transAxes,color=MUTED,fontsize=7)
        at.text(0.96,y,f"{e['amount']:,.0f} ج.م",transform=at.transAxes,color=RED,fontsize=9,fontweight='bold',ha='right')

    buf=io.BytesIO()
    plt.savefig(buf,format='PNG',dpi=150,bbox_inches='tight',facecolor=BG,edgecolor='none')
    buf.seek(0); plt.close('all'); return buf

# ── Telegram Handlers ─────────────────────────────
REPORT_KEYWORDS = {'تقرير','ملخص','احصائيات','إحصائيات','كشف','اجمالي','إجمالي','report'}

async def cmd_start(update,context):
    await update.message.reply_text(
        '👋 *أهلاً! أنا بوت المصاريف* 💰\n\n'
        '📝 *سجّل مصروف:*\n• صرفت 200 بنزين\n• اكل 150\n• فاتورة كهرباء 350\n\n'
        '📊 *اطلب تقرير:* ابعت كلمة *تقرير*\n\n'
        '_كل حاجة في Excel على OneDrive_ 🎉', parse_mode='Markdown')

async def handle_text(update,context):
    text=update.message.text.strip()
    if any(kw in text for kw in REPORT_KEYWORDS):
        await send_report(update,context); return
    msg=await update.message.reply_text('⏳ بحلل الرسالة...')
    try:
        parsed=parse_expense_ai(text)
        if not parsed.get('is_expense'):
            await msg.edit_text('❓ مش فاهمتها كمصروف\n\nجرب: "صرفت 200 بنزين"\nأو ابعت *"تقرير"*',parse_mode='Markdown')
            return
        await msg.edit_text('💾 بسجل في Excel...')
        total=add_expense(parsed)
        await msg.edit_text(
            f'✅ *تم التسجيل!*\n\n💰 *المبلغ:* {parsed["amount"]:,.0f} ج.م\n'
            f'📂 *الفئة:* {parsed["category"]}\n🏷️ *التفاصيل:* {parsed["subcategory"]}\n'
            f'📝 *ملاحظة:* {parsed["notes"]}\n📅 {datetime.now().strftime("%Y-%m-%d %H:%M")}\n\n'
            f'_📊 إجمالي العمليات: {total}_',parse_mode='Markdown')
    except json.JSONDecodeError:
        await msg.edit_text('❌ مش قادر أحلل الرسالة. حاول تاني.')
    except Exception as e:
        logger.error(f'Error: {e}'); await msg.edit_text('❌ حصل خطأ. حاول تاني.')

async def send_report(update,context):
    msg=await update.message.reply_text('📊 بجهز التقرير...')
    try:
        expenses=fetch_expenses()
        if not expenses:
            await msg.edit_text('📭 *مفيش مصاريف لسه!* ابعتلي مصروف أولاً 🙂',parse_mode='Markdown'); return
        await msg.edit_text('🎨 برسم التقرير...')
        img=build_report_image(expenses); now=datetime.now()
        data=[e for e in expenses if e['month']==now.strftime('%Y-%m')]
        total=sum(e['amount'] for e in data)
        ct=defaultdict(float)
        for e in data: ct[e['category']]+=e['amount']
        top=sorted(ct.items(),key=lambda x:x[1],reverse=True)[:5]
        lines='\n'.join(f'  • {c}: {v:,.0f} ج.م ({v/total*100:.0f}%)' for c,v in top) if total>0 else 'لا يوجد بيانات'
        caption=(f'📊 *تقرير {MONTHS_AR[now.month]} {now.year}*\n\n'
                 f'💰 *الإجمالي:* {total:,.0f} ج.م\n📝 *العمليات:* {len(data)}\n\n'
                 f'*🏆 أكبر الفئات:*\n{lines}')
        await msg.delete()
        await update.message.reply_photo(photo=img,caption=caption,parse_mode='Markdown')
    except Exception as e:
        logger.error(f'Report error: {e}'); await msg.edit_text('❌ حصل خطأ في التقرير.')

def main():
    missing=[v for v in ['TELEGRAM_TOKEN','GEMINI_API_KEY','AZURE_CLIENT_ID','MS_REFRESH_TOKEN'] if not os.getenv(v)]
    if missing: raise SystemExit(f'❌ Missing: {", ".join(missing)}')
    app=Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler('start',cmd_start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND,handle_text))
    logger.info('🚀 Bot running!'); app.run_polling(drop_pending_updates=True)

if __name__=='__main__': main()
